import os
from dotenv import load_dotenv
from datetime import date

load_dotenv()
import smtplib
from email.mime.text import MIMEText

class EmailNotifier:
    def __init__(self):
        self.email = os.getenv("EMAIL")
        self.password = os.getenv("PASSWORD")

    def send_reminder(self, to, company, role):
        subject = f"Follow up: {company} - {role}"
        body = f"Hey! It's been 7+ days since you applied to {company} for {role}. Time to follow up!"
        
        msg = MIMEText(body)
        msg['Subject'] = subject
        msg['From'] = self.email
        msg['To'] = to

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(self.email, self.password)
            server.sendmail(self.email, to, msg.as_string())
            print(f" Reminder sent for {company}!")

class Job:
    def __init__(self, company, role, applied_date, status="Applied", notes=""):
        self.company = company
        self.role = role
        self.applied_date = applied_date
        self.status = status
        self.notes = notes

    def days_since_applied(self):
        today = date.today()
        return (today - self.applied_date).days

    def __str__(self):
        return f"{self.company} | {self.role} | {self.status} | {self.days_since_applied()} days ago"


class JobTracker:
    def __init__(self):
        self.jobs = []

    def add_job(self, company, role, applied_date, notes=""):
        job = Job(company, role, applied_date, notes=notes)
        self.jobs.append(job)
        print(f" Added: {company} - {role}")

    def show_all(self):
        if not self.jobs:
            print("No applications yet!")
            return
        print("\n--- All Applications ---")
        for job in self.jobs:
            print(job)

    def check_reminders(self):
        print("\n--- Pending Reminders (7+ days) ---")
        found = False
        for job in self.jobs:
            if job.days_since_applied() >= 7 and job.status == "Applied":
                print(f" Follow up: {job.company} | {job.role}")
                found = True
        if not found:
            print("No pending reminders!")
            
    def save_to_excel(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        
        # Header
        ws.append(["Company", "Role", "Applied Date", "Status", "Days Since Applied", "Notes"])
        
        # Data
        for job in self.jobs:
            ws.append([
                job.company,
                job.role,
                str(job.applied_date),
                job.status,
                job.days_since_applied(),
                job.notes
            ])
        
        wb.save("job_applications.xlsx")
        print(" Excel file saved!")

    def load_from_excel(self):
        from openpyxl import load_workbook
        try:
            wb = load_workbook("job_applications.xlsx")
            ws = wb.active
            for row in list(ws.iter_rows(values_only=True))[1:]:  # Skip header
                company, role, applied_date, status, _, notes = row
                job = Job(company, role, date.fromisoformat(applied_date), status, notes or "")
                self.jobs.append(job)
            print("Data loaded from Excel!")
        except FileNotFoundError:
            print("No existing data found — starting fresh!")


if __name__ == "__main__":
    tracker = JobTracker()
    notifier = EmailNotifier()
    
    # Pehle load karo existing data
    tracker.load_from_excel()
    
    # Naye jobs add karo
    tracker.add_job("Google", "Python Intern", date(2026, 2, 20))
    tracker.add_job("Razorpay", "Backend Intern", date(2026, 2, 10))
    
    # Sab dikhao
    tracker.show_all()
    
    # Reminders check karo
    tracker.check_reminders()
    
    # Email bhejo
    for job in tracker.jobs:
        if job.days_since_applied() >= 7 and job.status == "Applied":
            notifier.send_reminder(notifier.email, job.company, job.role)
    
    # Excel mein save karo
    tracker.save_to_excel()